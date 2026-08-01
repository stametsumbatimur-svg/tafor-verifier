# 🔥 SIVETA - Excel Export Engine (BMKG Annex 3 / SOP/024/DM/X/2025)
import io
import re
from datetime import datetime
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import pandas as pd
from xlsxwriter.utility import xl_col_to_name

# ==========================================
# KAMUS PEMETAAN BULAN BAHASA INDONESIA
# ==========================================
BULAN_INDO = {
    1: "JANUARI",
    2: "FEBRUARI",
    3: "MARET",
    4: "APRIL",
    5: "MEI",
    6: "JUNI",
    7: "JULI",
    8: "AGUSTUS",
    9: "SEPTEMBER",
    10: "OKTOBER",
    11: "NOVEMBER",
    12: "DESEMBER",
    "1": "JANUARI",
    "2": "FEBRUARI",
    "3": "MARET",
    "4": "APRIL",
    "5": "MEI",
    "6": "JUNI",
    "7": "JULI",
    "8": "AGUSTUS",
    "9": "SEPTEMBER",
    "10": "OKTOBER",
    "11": "NOVEMBER",
    "12": "DESEMBER",
    "01": "JANUARI",
    "02": "FEBRUARI",
    "03": "MARET",
    "04": "APRIL",
    "05": "MEI",
    "06": "JUNI",
    "07": "JULI",
    "08": "AGUSTUS",
    "09": "SEPTEMBER",
    "10": "OKTOBER",
    "11": "NOVEMBER",
    "12": "DESEMBER",
    "JANUARY": "JANUARI",
    "FEBRUARY": "FEBRUARI",
    "MARCH": "MARET",
    "APRIL": "APRIL",
    "MAY": "MEI",
    "JUNE": "JUNI",
    "JULY": "JULI",
    "AUGUST": "AGUSTUS",
    "SEPTEMBER": "SEPTEMBER",
    "OCTOBER": "OKTOBER",
    "NOVEMBER": "NOVEMBER",
    "DECEMBER": "DESEMBER",
}


# ==============================================================================
# 1. EXPORT REKAP SINGLE SHEET ('VERIFIKASI TAFOR') - XLSXWRITER ENGINE
# ==============================================================================
def export_v_final_excel(
    df_vfinal,
    bulan,
    tahun,
    stasiun,
    nama_petugas,
    nip_petugas="[NIP PETUGAS]",
    nama_kepala="[NAMA KEPALA STASIUN]",
    nip_kepala="[NIP KEPALA]",
):
    df_excel = df_vfinal.copy()

    # Konversi input bulan ke Bahasa Indonesia
    bulan_str = BULAN_INDO.get(str(bulan).strip().upper(), str(bulan).upper())

    # Kolom standar yang wajib ada
    kolom_standar = [
        "Tanggal",
        "Jam_UTC",
        "Perubahan",
        "T_Arah",
        "T_Kec",
        "T_Vis",
        "T_Wx",
        "T_AwanJml",
        "T_AwanTgi",
        "M_Arah",
        "S_Arah",
        "M_Kec",
        "S_Kec",
        "M_Vis",
        "S_Vis",
        "M_Wx",
        "S_Wx",
        "M_AwanJml",
        "S_AwanJml",
        "M_AwanTgi",
        "S_AwanTgi",
    ]

    # Reindex kolom jika ada di dataframe untuk menjaga urutan data tetap presisi
    col_exist = [c for c in kolom_standar if c in df_excel.columns]
    if len(col_exist) > 0:
        df_excel = df_excel[col_exist]

    # Konversi Boolean / Angka Skor menjadi 'B' dan 'S'
    kolom_skor = ["S_Arah", "S_Kec", "S_Vis", "S_Wx", "S_AwanJml", "S_AwanTgi"]
    for col in kolom_skor:
        if col in df_excel.columns:
            df_excel[col] = df_excel[col].apply(
                lambda x: (
                    "S"
                    if str(x).strip().upper() in ["FALSE", "SALAH", "S", "0", ""]
                    else (
                        "B"
                        if str(x).strip().upper() in ["TRUE", "BENAR", "B", "1"]
                        else x
                    )
                )
            )

    # Kosongkan tampilan tanggal duplikat untuk keindahan visual (merge look)
    if "Tanggal" in df_excel.columns:
        df_excel.loc[df_excel["Tanggal"].duplicated(), "Tanggal"] = ""

    output = io.BytesIO()
    writer = pd.ExcelWriter(output, engine="xlsxwriter")
    workbook = writer.book
    worksheet = workbook.add_worksheet("VERIFIKASI TAFOR")

    # Format Styles
    format_title = workbook.add_format(
        {"bold": True, "align": "center", "valign": "vcenter", "font_size": 11}
    )
    format_subtitle = workbook.add_format(
        {"align": "center", "valign": "vcenter", "font_size": 9}
    )
    format_bold_left = workbook.add_format(
        {"bold": True, "align": "left", "valign": "vcenter", "font_size": 9.5}
    )

    format_req_header = workbook.add_format(
        {
            "border": 1,
            "bold": True,
            "align": "center",
            "valign": "vcenter",
            "bg_color": "#D9D9D9",
            "text_wrap": True,
            "font_size": 8.5,
        }
    )
    format_req_text = workbook.add_format(
        {
            "border": 1,
            "align": "left",
            "valign": "vcenter",
            "text_wrap": True,
            "font_size": 8,
        }
    )
    format_req_bold = workbook.add_format(
        {
            "border": 1,
            "bold": True,
            "align": "left",
            "valign": "vcenter",
            "font_size": 8.5,
        }
    )

    format_border_bold = workbook.add_format(
        {
            "border": 1,
            "bold": True,
            "align": "center",
            "valign": "vcenter",
            "text_wrap": True,
            "font_size": 8.5,
        }
    )
    format_persen = workbook.add_format(
        {
            "border": 1,
            "bold": True,
            "align": "center",
            "num_format": "0.00%",
            "font_size": 8.5,
        }
    )

    format_tabel_header = workbook.add_format(
        {
            "border": 1,
            "bold": True,
            "align": "center",
            "valign": "vcenter",
            "bg_color": "#D9D9D9",
            "text_wrap": True,
            "font_size": 8.5,
        }
    )
    format_tabel_data = workbook.add_format(
        {
            "border": 1,
            "align": "center",
            "valign": "vcenter",
            "font_size": 8.5,
            "text_wrap": True,
        }
    )

    format_hijau = workbook.add_format(
        {
            "bg_color": "#C6EFCE",
            "font_color": "#006100",
            "align": "center",
            "valign": "vcenter",
            "border": 1,
            "font_size": 8.5,
        }
    )
    format_merah = workbook.add_format(
        {
            "bg_color": "#FFC7CE",
            "font_color": "#9C0006",
            "align": "center",
            "valign": "vcenter",
            "border": 1,
            "font_size": 8.5,
        }
    )

    format_ttd_nama = workbook.add_format(
        {
            "bold": True,
            "align": "center",
            "valign": "vcenter",
            "font_size": 9.5,
            "underline": True,
        }
    )

    # Menentukan batas kolom
    max_col_data = len(df_excel.columns) - 1
    batas_col = max(max_col_data, 20)

    nama_kolom_cantik = [
        "Tgl",
        "Jangka Waktu",
        "Perubahan",
        "Arah\n(T)",
        "Kec\n(T)",
        "Vis\n(T)",
        "Cuaca\n(T)",
        "Jml\nAwan\n(T)",
        "Tgi\nAwan\n(T)",
        "Arah\n(M)",
        "Skor",
        "Kec\n(M)",
        "Skor",
        "Vis\n(M)",
        "Skor",
        "Cuaca\n(M)",
        "Skor",
        "Jml\nAwan\n(M)",
        "Skor",
        "Tgi\nAwan\n(M)",
        "Skor",
    ]

    # Menulis Header Tabel Manual
    worksheet.set_row(12, 38)
    for col_num, col_name in enumerate(
        nama_kolom_cantik[: len(df_excel.columns)]
    ):
        worksheet.write(12, col_num, col_name, format_tabel_header)

    # Menulis Data Rows
    for row_num, row_data in enumerate(df_excel.values):
        for col_num, value in enumerate(row_data):
            val = "" if pd.isna(value) else value
            worksheet.write(13 + row_num, col_num, val, format_tabel_data)

    # Header & Narasi SOP
    worksheet.merge_range(
        0, 0, 0, batas_col, "VERIFIKASI AERODROME FORECAST", format_title
    )
    worksheet.merge_range(
        1,
        0,
        1,
        batas_col,
        "Standard Operating Procedures (SOP) Nomor: SOP/024/DM/X/2025",
        format_title,
    )
    worksheet.write(
        3, 0, "PERSYARATAN / TOLERANSI KETELITIAN PRAKIRAAN :", format_bold_left
    )

    worksheet.merge_range(4, 0, 4, 1, "UNSUR METEOROLOGI", format_req_header)
    worksheet.merge_range(
        4, 2, 4, 7, "PERSYARATAN / TOLERANSI KETELITIAN", format_req_header
    )
    worksheet.merge_range(4, 8, 4, 10, "UNSUR METEOROLOGI", format_req_header)
    worksheet.merge_range(
        4,
        11,
        4,
        batas_col,
        "PERSYARATAN / TOLERANSI KETELITIAN",
        format_req_header,
    )

    worksheet.merge_range(5, 0, 5, 1, "A. Arah Angin", format_req_bold)
    worksheet.merge_range(
        5,
        2,
        5,
        7,
        "Benar apabila arah sama, atau selisih <= 60 derajat. Jika kecepatan"
        " angin <10 kt, atau VRB, atau kondisi CB/TS, dianggap benar.",
        format_req_text,
    )
    worksheet.merge_range(5, 8, 5, 10, "E. Jumlah Awan", format_req_bold)
    worksheet.merge_range(
        5,
        11,
        5,
        batas_col,
        "Benar apabila berada pada kelompok yang sama: FEW/SCT atau BKN/OVC."
        " Jika tinggi awan > 5000 ft, dianggap benar.",
        format_req_text,
    )

    worksheet.merge_range(6, 0, 6, 1, "B. Kecepatan Angin", format_req_bold)
    worksheet.merge_range(
        6,
        2,
        6,
        7,
        "Selisih kecepatan dasar <= 10 knot. Status gust harus konsisten.",
        format_req_text,
    )
    worksheet.merge_range(6, 8, 6, 10, "F. Tinggi Dasar Awan", format_req_bold)
    worksheet.merge_range(
        6,
        11,
        6,
        batas_col,
        "Selisih <= 100 ft untuk <1000 ft. Untuk >= 1000 ft, selisih <= 30%"
        " dari tinggi awan Manual.",
        format_req_text,
    )

    worksheet.merge_range(7, 0, 7, 1, "C. Jarak Pandang", format_req_bold)
    worksheet.merge_range(
        7,
        2,
        7,
        7,
        "Benar apabila berada pada kelas visibility yang sama.",
        format_req_text,
    )
    worksheet.merge_range(7, 8, 7, 10, "", format_req_bold)
    worksheet.merge_range(7, 11, 7, batas_col, "", format_req_text)

    worksheet.merge_range(8, 0, 8, 1, "D. Cuaca / Endapan", format_req_bold)
    worksheet.merge_range(
        8,
        2,
        8,
        7,
        "Benar apabila sama-sama mendeteksi atau tidak mendeteksi presipitasi"
        " sedang/lebat. Hujan ringan (-RA) tidak dihitung.",
        format_req_text,
    )
    worksheet.merge_range(8, 8, 8, 10, "", format_req_bold)
    worksheet.merge_range(8, 11, 8, batas_col, "", format_req_text)

    worksheet.set_row(5, 68)
    worksheet.set_row(6, 52)
    worksheet.set_row(7, 35)
    worksheet.set_row(8, 68)

    worksheet.write(10, 0, f"BULAN : {bulan_str}", format_bold_left)
    worksheet.write(10, 3, f"TAHUN : {tahun}", format_bold_left)
    worksheet.write(10, 6, "(SEMUA WAKTU DALAM GMT)", format_bold_left)
    worksheet.write(10, 11, f"STASIUN METEOROLOGI {stasiun}", format_bold_left)

    # Conditional Formatting B/S
    jumlah_baris_data = len(df_excel)
    excel_start_data_row = 14
    excel_last_data_row = excel_start_data_row + jumlah_baris_data - 1

    data_range = f"A{excel_start_data_row}:{xl_col_to_name(max_col_data)}{excel_last_data_row}"
    worksheet.conditional_format(
        data_range,
        {
            "type": "cell",
            "criteria": "==",
            "value": '"B"',
            "format": format_hijau,
        },
    )
    worksheet.conditional_format(
        data_range,
        {
            "type": "cell",
            "criteria": "==",
            "value": '"S"',
            "format": format_merah,
        },
    )

    worksheet.autofilter(12, 0, 12 + jumlah_baris_data, max_col_data)
    worksheet.freeze_panes(13, 0)

    # Footer Summaries (JUMLAH & PROSENTASE KEBENARAN)
    baris_jumlah_idx = 12 + jumlah_baris_data + 1
    excel_baris_jumlah = baris_jumlah_idx + 1
    baris_persen_idx = baris_jumlah_idx + 1

    worksheet.merge_range(
        baris_jumlah_idx, 0, baris_jumlah_idx, 2, "JUMLAH", format_border_bold
    )
    worksheet.merge_range(
        baris_persen_idx,
        0,
        baris_persen_idx,
        2,
        "PROSENTASE KEBENARAN",
        format_border_bold,
    )

    # Isi nilai default sampel non-skor
    for col_idx in range(3, max_col_data + 1):
        worksheet.write(
            baris_jumlah_idx, col_idx, jumlah_baris_data, format_border_bold
        )
        worksheet.write(baris_persen_idx, col_idx, "", format_border_bold)

    # Overwrite rumus pada kolom-kolom skor B/S
    for col_name in kolom_skor:
        if col_name in df_excel.columns:
            col_idx = df_excel.columns.get_loc(col_name)
            col_huruf = xl_col_to_name(col_idx)

            # Hitung sampel B + S saja (mengabaikan NIL / -)
            rumus_jumlah = (
                f'=COUNTIF({col_huruf}{excel_start_data_row}:{col_huruf}{excel_last_data_row},'
                f' "B") +'
                f' COUNTIF({col_huruf}{excel_start_data_row}:{col_huruf}{excel_last_data_row},'
                ' "S")'
            )
            worksheet.write_formula(
                baris_jumlah_idx, col_idx, rumus_jumlah, format_border_bold
            )

            # Persentase Kebenaran = B / Total (B+S)
            rumus_persen = (
                f'=IFERROR(COUNTIF({col_huruf}{excel_start_data_row}:{col_huruf}{excel_last_data_row},'
                f' "B") / {col_huruf}{excel_baris_jumlah}, 0)'
            )
            worksheet.write_formula(
                baris_persen_idx, col_idx, rumus_persen, format_persen
            )

    # Tanda Tangan
    baris_ttd = baris_persen_idx + 4
    worksheet.merge_range(
        baris_ttd, 0, baris_ttd, 3, "Mengetahui,", format_subtitle
    )
    worksheet.merge_range(
        baris_ttd + 1, 0, baris_ttd + 1, 3, "Kepala Stasiun", format_subtitle
    )
    worksheet.merge_range(
        baris_ttd + 5, 0, baris_ttd + 5, 3, nama_kepala, format_ttd_nama
    )
    worksheet.merge_range(
        baris_ttd + 6,
        0,
        baris_ttd + 6,
        3,
        f"NIP. {nip_kepala}",
        format_subtitle,
    )

    col_ttd_start = max(batas_col - 4, 4)
    col_ttd_end = batas_col
    worksheet.merge_range(
        baris_ttd,
        col_ttd_start,
        baris_ttd,
        col_ttd_end,
        "Petugas Pembuat Laporan",
        format_subtitle,
    )
    worksheet.merge_range(
        baris_ttd + 5,
        col_ttd_start,
        baris_ttd + 5,
        col_ttd_end,
        nama_petugas,
        format_ttd_nama,
    )
    worksheet.merge_range(
        baris_ttd + 6,
        col_ttd_start,
        baris_ttd + 6,
        col_ttd_end,
        f"NIP. {nip_petugas}",
        format_subtitle,
    )

    # Format Lebar Kolom
    worksheet.set_column("A:A", 4.5)
    worksheet.set_column("B:C", 9.5)
    worksheet.set_column("D:E", 6.0)
    worksheet.set_column("F:I", 7.0)
    worksheet.set_column("J:J", 6.0)
    worksheet.set_column("K:K", 6.8)
    worksheet.set_column("L:L", 6.0)
    worksheet.set_column("M:M", 6.8)
    worksheet.set_column("N:N", 7.0)
    worksheet.set_column("O:O", 6.8)
    worksheet.set_column("P:P", 7.0)
    worksheet.set_column("Q:Q", 6.8)
    worksheet.set_column("R:R", 7.0)
    worksheet.set_column("S:S", 6.8)
    worksheet.set_column("T:T", 7.0)
    worksheet.set_column("U:U", 6.8)

    # Page Print Setup
    worksheet.set_portrait()
    worksheet.set_paper(9)  # A4
    worksheet.fit_to_pages(1, 0)
    worksheet.set_margins(left=0.15, right=0.15, top=0.4, bottom=0.4)
    worksheet.repeat_rows(12, 12)

    akhir_baris_print = baris_ttd + 7
    worksheet.print_area(0, 0, akhir_baris_print, max_col_data)

    writer.close()
    return output.getvalue()


# ==============================================================================
# 2. EXPORT MULTI-SHEET REKAP 1 BULAN + TANGGAL 1-31 (OPENPYXL ENGINE)
# ==============================================================================
def generate_klasik_31_sheet(df_filtered):
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Hapus sheet default

    # Tentukan Bulan dan Tahun dari dataset
    nama_bulan = "JUNI"
    tahun = "2026"

    if not df_filtered.empty and "Waktu Aktual (UTC)" in df_filtered.columns:
        try:
            contoh_waktu = pd.to_datetime(
                df_filtered.iloc[0]["Waktu Aktual (UTC)"]
            )
            nama_bulan = BULAN_INDO.get(contoh_waktu.month, "JUNI")
            tahun = str(contoh_waktu.year)
        except Exception:
            pass

    font_bold = Font(name="Calibri", size=9, bold=True)
    font_normal = Font(name="Calibri", size=9)

    # --------------------------------------------------------------------------
    # A. SHEET 'rekap 1 bulan'
    # --------------------------------------------------------------------------
    ws_rekap = wb.create_sheet(title="rekap 1 bulan")

    ws_rekap.cell(
        row=1,
        column=1,
        value=(
            "**Catatan untuk verifikator: form ini disediakan sebagai format"
            " pelaporan. jika terdapat error dalam penghitungan, formula boleh"
            " dikoreksi dengan mengikuti aturan dalam lampiran surat Deputi"
            " Bidang Meteorologi ME.00.00/026/DM/XII/2020"
        ),
    )
    ws_rekap.cell(
        row=2, column=1, value=f"VERIFIKASI TAF BULAN {nama_bulan} {tahun}"
    ).font = Font(name="Calibri", size=11, bold=True)

    headers_unsur = [
        "unsur",
        None,
        "A. Arah angin",
        None,
        "B. Kecepatan Angin",
        None,
        "B.2 GUSTY",
        None,
        "C. Jarak Pandang",
        None,
        "D. Endapan",
        None,
        "E. Jumlah Awan",
        None,
        "F. Ketinggian Awan",
    ]

    for c_idx, val in enumerate(headers_unsur, 1):
        if val:
            ws_rekap.cell(row=3, column=c_idx, value=val).font = font_bold
            ws_rekap.cell(row=5, column=c_idx, value=val).font = font_bold

    # Rata-rata 1 Bulan (Formula)
    ws_rekap.cell(row=4, column=1, value="rata2 1 bulan").font = font_bold
    ws_rekap.cell(row=4, column=3, value="=AVERAGE(C7:C130)")
    ws_rekap.cell(row=4, column=5, value="=AVERAGE(E7:E130)")
    ws_rekap.cell(row=4, column=7, value="=AVERAGE(G7:G130)")
    ws_rekap.cell(row=4, column=9, value="=AVERAGE(I7:I130)")
    ws_rekap.cell(row=4, column=11, value="=AVERAGE(K7:K130)")
    ws_rekap.cell(row=4, column=13, value="=AVERAGE(M7:M130)")
    ws_rekap.cell(row=4, column=15, value="=AVERAGE(O7:O130)")

    ws_rekap.cell(row=6, column=1, value="tanggal/taf").font = font_bold
    ws_rekap.cell(row=6, column=2, value="jam taf").font = font_bold

    # Isi Tabel Rekap 31 Hari x 4 Jam TAF (00, 06, 12, 18 UTC)
    rekap_r = 7
    for hari in range(1, 32):
        day_str = str(hari)
        for jam_idx, jam_str in enumerate(["00", "06", "12", "18"]):
            ws_rekap.cell(
                row=rekap_r, column=1, value=hari if jam_idx == 0 else None
            )
            ws_rekap.cell(row=rekap_r, column=2, value=jam_str)

            # Tentukan Baris Hasil Verifikasi pada Sheet Harian
            target_res_row = (
                65
                if jam_str == "00"
                else (131 if jam_str == "06" else (196 if jam_str == "12" else 262))
            )

            ws_rekap.cell(
                row=rekap_r, column=3, value=f"='{day_str}'!O${target_res_row}"
            )
            ws_rekap.cell(
                row=rekap_r, column=5, value=f"='{day_str}'!Q${target_res_row}"
            )
            ws_rekap.cell(
                row=rekap_r, column=7, value=f"='{day_str}'!S${target_res_row}"
            )
            ws_rekap.cell(
                row=rekap_r, column=9, value=f"='{day_str}'!U${target_res_row}"
            )
            ws_rekap.cell(
                row=rekap_r, column=11, value=f"='{day_str}'!W${target_res_row}"
            )
            ws_rekap.cell(
                row=rekap_r, column=13, value=f"='{day_str}'!Y${target_res_row}"
            )
            ws_rekap.cell(
                row=rekap_r,
                column=15,
                value=f"='{day_str}'!AA${target_res_row}",
            )

            rekap_r += 1

    # --------------------------------------------------------------------------
    # B. SHEET HARIAN ('1' S.D. '31')
    # --------------------------------------------------------------------------
    for hari in range(1, 32):
        day_str = str(hari)
        ws_day = wb.create_sheet(title=day_str)

        ws_day.cell(
            row=1,
            column=2,
            value=(
                "**Catatan untuk verifikator: form ini disediakan sebagai"
                " format pelaporan."
            ),
        )
        ws_day.cell(row=3, column=3, value=f"BULAN : {nama_bulan}")
        ws_day.cell(row=3, column=5, value=f"TAHUN : {tahun}")
        ws_day.cell(row=3, column=9, value="( SEMUA WAKTU DALAM UTC )")

        curr_r = 4
        for jam_taf in ["00", "06", "12", "18"]:
            ws_day.cell(row=curr_r, column=2, value="Tanggal ")
            ws_day.cell(row=curr_r, column=3, value="Jangka waktu")
            ws_day.cell(row=curr_r, column=4, value="Prakiraan : ")
            ws_day.cell(
                row=curr_r, column=13, value="KENYATAAN (METAR DAN SPECI):"
            )

            curr_r += 1
            ws_day.cell(row=curr_r, column=4, value="Change Group")
            ws_day.cell(row=curr_r, column=5, value="Change Group Time (UTC)")

            cols_sub = [
                "A",
                "B1",
                "B2",
                "C",
                "D",
                "E",
                "F",
                "DATA METAR",
                "A",
                "H",
                "B1",
                "H",
                "B2",
                "H",
                "C",
                "H",
                "D",
                "H",
                "E",
                "H",
                "F",
                "H",
            ]
            for idx_c, col_title in enumerate(cols_sub, 6):
                ws_day.cell(row=curr_r, column=idx_c, value=col_title)

            curr_r += 2  # Baris data dimulai dari row 7, 73, 139, 204
            start_data_r = curr_r

            # Tulis 48 Interval Observasi
            for obs_idx in range(1, 49):
                r = curr_r
                ws_day.cell(row=r, column=1, value=obs_idx)
                ws_day.cell(
                    row=r, column=2, value=hari if obs_idx == 1 else None
                )
                ws_day.cell(row=r, column=4, value="GENERAL")

                # Rumus Parsial METAR & Evaluasi Toleransi ICAO / BMKG Pusat
                ws_day.cell(
                    row=r,
                    column=14,
                    value=(
                        f'=IF(ISNUMBER(VALUE(FIND("VRB",M{r}))),999,VALUE(MID(M{r},20,3)))'
                    ),
                )
                ws_day.cell(
                    row=r,
                    column=15,
                    value=(
                        f"=IF(OR(ABS(F{r}-N{r})<61,AND(ABS(F{r}-N{r})>60,P{r}<10),F{r}=N{r}),1,0)"
                    ),
                )
                ws_day.cell(
                    row=r, column=16, value=f"=VALUE(MID(M{r},23,2))"
                )
                ws_day.cell(
                    row=r, column=17, value=f"=IF(ABS(P{r}-G{r})<11,1,0)"
                )
                ws_day.cell(
                    row=r,
                    column=18,
                    value=f'=IF(MID(M{r},25,1)="G","GUST",0)',
                )
                ws_day.cell(row=r, column=19, value=f"=IF(H{r}=R{r},1,0)")
                ws_day.cell(
                    row=r,
                    column=20,
                    value=(
                        f'=IF(ISNUMBER(VALUE(MID(M{r},(FIND("KT",M{r})+3),4))),VALUE(MID(M{r},(FIND("KT",M{r})+3),4)),9999)'
                    ),
                )
                ws_day.cell(
                    row=r,
                    column=21,
                    value=(
                        f"=IF(OR(AND(I{r}<801,T{r}<801),AND(I{r}>799,I{r}<1501,T{r}>799,T{r}<1502),AND(I{r}>1499,I{r}<3001,T{r}>1499,T{r}<3001),AND(I{r}>2999,I{r}<5001,T{r}>2999,T{r}<5001),AND(I{r}>4999,T{r}>4999)),1,0)"
                    ),
                )
                ws_day.cell(
                    row=r,
                    column=22,
                    value=(
                        f'=IF(AND(IFERROR(FIND("RA",M{r}),0)>0,IFERROR(FIND("RA",M{r}),0)<38),"RA",0)'
                    ),
                )
                ws_day.cell(row=r, column=23, value=f"=IF(EXACT(V{r},J{r}),1,0)")
                ws_day.cell(
                    row=r,
                    column=24,
                    value=(
                        f'=IF(ISNUMBER(VALUE(FIND("FEW",M{r}))),"FEW",IF(ISNUMBER(VALUE(FIND("SCT",M{r}))),"SCT",IF(ISNUMBER(VALUE(FIND("BKN",M{r}))),"BKN",IF(ISNUMBER(VALUE(FIND("OVC",M{r}))),"OVC",IF(ISNUMBER(VALUE(FIND("SKC",M{r}))),"SKC","SKC")))))'
                    ),
                )
                ws_day.cell(
                    row=r,
                    column=25,
                    value=(
                        f'=IF(OR(AND(OR(K{r}="SKC",K{r}="FEW",K{r}="SCT"),OR(X{r}="SKC",X{r}="FEW",X{r}="SCT")),(AND(OR(K{r}="BKN",K{r}="OVC"),OR(X{r}="BKN",X{r}="OVC")))),1,0)'
                    ),
                )
                ws_day.cell(
                    row=r,
                    column=26,
                    value=f'=IF(ISNUMBER(VALUE(FIND("CAVOK",M{r}))),9999,1800)',
                )
                ws_day.cell(
                    row=r,
                    column=27,
                    value=(
                        f"=IF(AND(Z{r}<1001,ABS(L{r}-Z{r})<101),1,IF(AND(Z{r}>1000,ABS(Z{r}-L{r})/L{r}<0.30005),1,0))"
                    ),
                )

                curr_r += 1

            end_data_r = curr_r - 1

            # Baris Agregasi Summary (12 JAM, UMUM, BECMG, TEMPO, hasil verifikasi)
            curr_r += 6
            ws_day.cell(row=curr_r, column=4, value="12 JAM")
            for c_col, letter in [
                (15, "O"),
                (17, "Q"),
                (19, "S"),
                (21, "U"),
                (23, "W"),
                (25, "Y"),
                (27, "AA"),
            ]:
                ws_day.cell(
                    row=curr_r,
                    column=c_col,
                    value=f"=SUM({letter}{start_data_r}:{letter}{end_data_r})",
                )

            curr_r += 1  # UMUM
            ws_day.cell(row=curr_r, column=4, value="UMUM")
            for c_col, letter in [
                (15, "O"),
                (17, "Q"),
                (19, "S"),
                (21, "U"),
                (23, "W"),
                (25, "Y"),
                (27, "AA"),
            ]:
                ws_day.cell(
                    row=curr_r,
                    column=c_col,
                    value=(
                        f'=SUMIF($D{start_data_r}:$D{end_data_r},"GENERAL",{letter}{start_data_r}:{letter}{end_data_r})'
                    ),
                )

            curr_r += 1  # BECMG
            ws_day.cell(row=curr_r, column=4, value="BECMG")
            for c_col, letter in [
                (15, "O"),
                (17, "Q"),
                (19, "S"),
                (21, "U"),
                (23, "W"),
                (25, "Y"),
                (27, "AA"),
            ]:
                ws_day.cell(
                    row=curr_r,
                    column=c_col,
                    value=(
                        f'=SUMIF($D{start_data_r}:$D{end_data_r},"BECMG",{letter}{start_data_r}:{letter}{end_data_r})'
                    ),
                )

            curr_r += 2  # TEMPO 1
            ws_day.cell(row=curr_r, column=4, value="TEMPO 1")

            curr_r += 1  # hasil verifikasi
            ws_day.cell(row=curr_r, column=4, value="hasil verifikasi")

            gen_r = curr_r - 4
            tempo_r = curr_r - 1

            for c_col, letter in [
                (15, "O"),
                (17, "Q"),
                (19, "S"),
                (21, "U"),
                (23, "W"),
                (25, "Y"),
                (27, "AA"),
            ]:
                ws_day.cell(
                    row=curr_r,
                    column=c_col,
                    value=(
                        f"=(SUM({letter}{gen_r}:{letter}{tempo_r})/COUNTA({letter}{start_data_r}:{letter}{end_data_r}))*100"
                    ),
                )

            curr_r += 2  # Persiapan ke blok TAF berikutnya

    wb.save(output)
    output.seek(0)
    return output
